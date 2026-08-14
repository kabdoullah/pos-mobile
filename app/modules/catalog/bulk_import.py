"""Parsing des fichiers d'import en masse (CSV / Excel).

Chaque ligne de données est traitée indépendamment : une ligne malformée (valeur
numérique invalide, champ requis manquant) ne fait jamais échouer les autres —
contrairement à l'endpoint JSON `POST /products/bulk`, la donnée source ici vient
d'un système tiers non maîtrisé, pas d'un opérateur technique préparant du JSON propre.
"""

import csv
import io
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError as PydanticValidationError

from app.core.exceptions import ValidationError
from app.modules.catalog.schemas import ProductCreate

_MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024
_MAX_ROWS = 500

TEMPLATE_COLUMNS = ("name", "barcode", "unit_price", "current_stock", "min_stock")
_TEMPLATE_ROWS = (
    ("Riz local 5kg", "", "3500.00", "50", "10"),
    ("Savon Zest", "6191234567890", "500.00", "120", "20"),
)

# Ligne de données : (index, produit si valide, message d'erreur si invalide)
ParsedRow = tuple[int, ProductCreate | None, str | None]


def parse_bulk_import_file(filename: str, content: bytes) -> list[ParsedRow]:
    """Parse un fichier CSV ou Excel (.xlsx) en lignes ProductCreate.

    Lève ValidationError (422) uniquement pour des problèmes globaux au fichier
    (extension non supportée, taille, nombre de lignes) — jamais pour une ligne
    de données individuelle, qui est simplement marquée en échec.
    """
    if len(content) > _MAX_FILE_SIZE_BYTES:
        raise ValidationError("File too large (max 5 MB).", field="file")

    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension == "csv":
        rows = _read_csv_rows(content)
    elif extension == "xlsx":
        rows = _read_xlsx_rows(content)
    else:
        raise ValidationError("Unsupported file type. Use .csv or .xlsx.", field="file")

    if len(rows) > _MAX_ROWS:
        raise ValidationError(f"Too many rows (max {_MAX_ROWS}).", field="file")

    return [_row_to_product(index, row) for index, row in enumerate(rows)]


def _decode_csv_bytes(content: bytes) -> str:
    """utf-8-sig gère le BOM ajouté par Excel ; cp1252 en repli pour les accents
    des exports Excel FR qui n'utilisent pas UTF-8."""
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return content.decode("cp1252")


def _read_csv_rows(content: bytes) -> list[dict[str, object]]:
    text = _decode_csv_bytes(content)
    sample = text[:2048]
    try:
        dialect: type[csv.Dialect] | csv.Dialect = csv.Sniffer().sniff(sample, delimiters=",;")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return [
        {(key or "").strip().lower(): (value or "").strip() for key, value in row.items()}
        for row in reader
    ]


def _read_xlsx_rows(content: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    normalized_header = [str(cell or "").strip().lower() for cell in header]

    rows: list[dict[str, object]] = []
    for raw_row in rows_iter:
        if all(cell is None for cell in raw_row):
            continue
        rows.append(dict(zip(normalized_header, raw_row, strict=False)))
    return rows


def _clean_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_decimal(value: object) -> Decimal:
    text = _clean_str(value)
    if not text:
        raise ValueError("unit_price is required")
    return Decimal(text.replace(",", "."))


def _parse_int(value: object) -> int | None:
    text = _clean_str(value)
    if not text:
        return None
    return int(float(text.replace(",", ".")))


def _row_to_product(index: int, row: dict[str, object]) -> ParsedRow:
    try:
        payload = {
            "name": _clean_str(row.get("name")),
            "barcode": _clean_str(row.get("barcode")) or None,
            "unit_price": _parse_decimal(row.get("unit_price")),
            "current_stock": _parse_int(row.get("current_stock")),
            "min_stock": _parse_int(row.get("min_stock")),
        }
    except (InvalidOperation, ValueError) as exc:
        return index, None, f"Invalid numeric value: {exc}"

    try:
        product = ProductCreate.model_validate(payload)
    except PydanticValidationError as exc:
        first_error = exc.errors()[0]
        field = first_error["loc"][0] if first_error.get("loc") else "unknown"
        return index, None, f"{field}: {first_error['msg']}"

    return index, product, None


def build_csv_template() -> bytes:
    """Génère le modèle CSV téléchargeable (`GET /products/bulk/template`).

    Encodé en utf-8-sig (BOM) pour qu'Excel l'ouvre correctement avec les accents.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(TEMPLATE_COLUMNS)
    writer.writerows(_TEMPLATE_ROWS)
    return buffer.getvalue().encode("utf-8-sig")


def build_xlsx_template() -> bytes:
    """Génère le modèle Excel téléchargeable (`GET /products/bulk/template`)."""
    workbook = Workbook()
    sheet = workbook.worksheets[0]
    sheet.append(TEMPLATE_COLUMNS)
    for row in _TEMPLATE_ROWS:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
